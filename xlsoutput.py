import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os

from models import JobListing

#column indexes
NUMERO = 0
TITULO_POSICION = 1
AREA = 2
ESPECIALIDAD = 3
DESCRIPCION_POSICION = 4
IDIOMA_1 = 5
NIVEL_IDIOMA_1 = 6
IDIOMA_2 = 7
NIVEL_IDIOMA_2 = 8
IDIOMA_3 = 9
NIVEL_IDIOMA_3 = 10
PAIS = 11
ESTADO = 12
CIUDAD = 13
TIPO_CONTRATACION = 14
JORNADA_LABORAL = 15
ANIOS_EXPERIENCIA = 16
MONEDA = 17
RANGO_SALARIAL = 18
MOSTRAR_SALARIO = 19
VIGENCIA_DIAS = 20
NOMBRE_EMPRESA = 21
DESCRIPCION_EMPRESA = 22
URL_ORIGINAL = 23

DEFAULT_FILENAME = f'vacantes.xls'

def openXlsxFile(filename):
    try:
        _ = open(filename)
        return readXlsxFile(filename)
    except IOError:
        return createXlsxFile(filename)
    
def readXlsxFile(filename):
    book = openpyxl.load_workbook(filename)
    return book


def createXlsxFile(filename):
    book = Workbook()
    sheet = book.active

    # Write the dumbass header
    sheet.cell(row=1, column=NUMERO).value = '#'
    sheet.cell(row=1, column=TITULO_POSICION).value='Titulo de la posición'
    sheet.cell(row=1, column=AREA).value='Área'
    sheet.cell(row=1, column=ESPECIALIDAD).value='Especialidad'
    sheet.cell(row=1, column=DESCRIPCION_POSICION).value='Descripción de la posición'
    sheet.cell(row=1, column=IDIOMA_1).value='Idioma'
    sheet.cell(row=1, column=NIVEL_IDIOMA_1).value='Nivel'
    sheet.cell(row=1, column=IDIOMA_2).value='Idioma'
    sheet.cell(row=1, column=NIVEL_IDIOMA_2).value='Nivel'
    sheet.cell(row=1, column=IDIOMA_3).value='Idioma'
    sheet.cell(row=1, column=NIVEL_IDIOMA_3).value='Nivel'
    sheet.cell(row=1, column=PAIS).value='País'
    sheet.cell(row=1, column=ESTADO).value='Estado'
    sheet.cell(row=1, column=CIUDAD).value='Ciudad'
    sheet.cell(row=1, column=TIPO_CONTRATACION).value='Tipo de contratación'
    sheet.cell(row=1, column=JORNADA_LABORAL).value='Jornada laboral'
    sheet.cell(row=1, column=ANIOS_EXPERIENCIA).value='Años de experiencia'
    sheet.cell(row=1, column=MONEDA).value='Moneda'
    sheet.cell(row=1, column=RANGO_SALARIAL).value='Rango salarial'
    sheet.cell(row=1, column=MOSTRAR_SALARIO).value='Mostrar salario en anuncio'
    sheet.cell(row=1, column=VIGENCIA_DIAS).value='Vigencia (días)'
    sheet.cell(row=1, column=NOMBRE_EMPRESA).value='Nombre empresa'
    sheet.cell(row=1, column=DESCRIPCION_EMPRESA).value='Descripción empresa'
    sheet.cell(row=1, column=URL_ORIGINAL).value='URL original'

    return book

def writeOnXlsxFile(row, column, field, workbook):
    sheet = workbook.active
    sheet.cell(row=row, column=column).value = field


def closeXlsxFile(file):
    file.close()

def insertJobs(jobs):
    workbook = openXlsxFile(DEFAULT_FILENAME)
    sheet = workbook.active
    row = len(sheet[get_column_letter(URL_ORIGINAL)]) + 1

    for job in jobs:
        writeOnXlsxFile(row, CIUDAD, job.city, workbook)
        writeOnXlsxFile(row, ESTADO, job.state, workbook)
        writeOnXlsxFile(row, TITULO_POSICION, job.title, workbook)
        writeOnXlsxFile(row, RANGO_SALARIAL, job.salaryString, workbook)
        writeOnXlsxFile(row, URL_ORIGINAL, job.url, workbook)

    workbook.save(DEFAULT_FILENAME)

def insertJob(job):
    workbook = openXlsxFile(DEFAULT_FILENAME)
    sheet = workbook.active
    row = len(sheet[get_column_letter(URL_ORIGINAL)]) + 1
    
    writeOnXlsxFile(row, CIUDAD, job.city, workbook)
    writeOnXlsxFile(row, ESTADO, job.state, workbook)
    writeOnXlsxFile(row, TITULO_POSICION, job.title, workbook)
    writeOnXlsxFile(row, RANGO_SALARIAL, job.salaryString, workbook)
    writeOnXlsxFile(row, URL_ORIGINAL, job.url, workbook)

    workbook.save(DEFAULT_FILENAME)




        